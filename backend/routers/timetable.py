import re
from typing import List
from pydantic import BaseModel
from datetime import datetime, timedelta
import io
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, Form, File
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from database import get_db
import models
from models import Timetable, Module, Lecturer
from utils.audit_logger import log_audit_action

router = APIRouter(prefix="/api/timetable", tags=["Timetable"])


#Helper: ordinal date suffix 
def _ordinal(n: int) -> str:
    if 11 <= (n % 100) <= 13:
        return f"{n}th"
    return f"{n}{['th','st','nd','rd','th'][min(n % 10, 4)]}"


def _fmt_date(date_str: str) -> str:
    """Convert '2026-04-22' → '22nd April 2026'."""
    try:
        d = datetime.strptime(str(date_str).strip(), "%Y-%m-%d")
        return f"{_ordinal(d.day)} {d.strftime('%B %Y')}"
    except Exception:
        return str(date_str)


def _parse_time(t: str) -> datetime | None:
    """Try common time formats and return a datetime (date-agnostic)."""
    for fmt in ("%I:%M %p", "%H:%M", "%I:%M%p", "%H:%M:%S"):
        try:
            return datetime.strptime(t.strip().upper(), fmt)
        except ValueError:
            continue
    return None


def _duration_hours(start: str, end: str) -> float:
    """Return decimal hours between start and end time strings."""
    s, e = _parse_time(start), _parse_time(end)
    if not s or not e:
        return 0.0
    delta = (e - s).total_seconds()
    return round(delta / 3600, 2) if delta > 0 else 0.0


def _fmt_duration(hours: float) -> str:
    total_mins = int(hours * 60)
    h, m = divmod(total_mins, 60)
    if m == 0:
        return f"{h} Hour{'s' if h != 1 else ''}"
    return f"{h}h {m}m"


def _session_status(date_str: str, start: str, end: str) -> str:
    """Classify as Completed / Ongoing / Upcoming relative to right now."""
    try:
        day = datetime.strptime(str(date_str).strip(), "%Y-%m-%d").date()
        s = _parse_time(start)
        e = _parse_time(end)
        if not s or not e:
            return "Unknown"
        session_start = datetime.combine(day, s.time())
        session_end   = datetime.combine(day, e.time())
        now = datetime.now()
        if now > session_end:
            return "Completed"
        if now >= session_start:
            return "Ongoing"
        return "Upcoming"
    except Exception:
        return "Unknown"


# Openpyxl style constants 
_DARK_BLUE   = "1F3864"
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_TITLE_FONT  = Font(name="Calibri", bold=True, size=14, color="1F3864")
_META_FONT   = Font(name="Calibri", italic=True, size=10, color="555555")
_HEADER_FILL = PatternFill("solid", fgColor=_DARK_BLUE)
_ALT_FILL    = PatternFill("solid", fgColor="EEF2FF")   # soft indigo stripe
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_THIN  = Side(style="thin", color="CCCCCC")
_THICK = Side(style="medium", color="1F3864")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_BORDER = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)

_STATUS_FILLS = {
    "Completed": PatternFill("solid", fgColor="D1FAE5"),  # green
    "Ongoing":   PatternFill("solid", fgColor="FEF9C3"),  # yellow
    "Upcoming":  PatternFill("solid", fgColor="DBEAFE"),  # blue
    "Unknown":   PatternFill("solid", fgColor="F3F4F6"),
}
_STATUS_FONTS = {
    "Completed": Font(name="Calibri", color="065F46", bold=True, size=10),
    "Ongoing":   Font(name="Calibri", color="92400E", bold=True, size=10),
    "Upcoming":  Font(name="Calibri", color="1E40AF", bold=True, size=10),
    "Unknown":   Font(name="Calibri", color="6B7280", size=10),
}


@router.get("/export/{batch_id}")
async def export_timetable_excel(batch_id: str, db: Session = Depends(get_db)):
    """
    Generates a professional, multi-sheet Excel report for a given batch.
    Sheet 1: Official Timetable with metadata, computed columns, and styling.
    Sheet 2: Analytics Summary with per-lecturer hours and session counts.
    """
    records = db.query(Timetable).filter(
        Timetable.batch_id == batch_id
    ).order_by(Timetable.date.asc(), Timetable.start_time.asc()).all()

    if not records:
        raise HTTPException(status_code=404, detail=f"No timetable records found for batch '{batch_id}'.")

    #  Pre-compute row data 
    rows = []
    for rec in records:
        dur_h = _duration_hours(rec.start_time or "", rec.end_time or "")
        rows.append({
            "date":        rec.date or "",
            "day":         getattr(rec, "day", "") or "",
            "start_time":  rec.start_time or "",
            "end_time":    rec.end_time or "",
            "module_code": rec.module_code or "",
            "module_name": getattr(rec, "module_name", "") or rec.module_code or "",
            "lecturer":    rec.lecturer or "TBA",
            "location":    rec.location or "TBA",
            "date_fmt":    _fmt_date(rec.date or ""),
            "duration_h":  dur_h,
            "duration":    _fmt_duration(dur_h),
            "status":      _session_status(rec.date or "", rec.start_time or "", rec.end_time or ""),
        })

    #  Build workbook 
    wb = openpyxl.Workbook()

    #  SHEET 1 — Official Timetable 
    ws = wb.active
    ws.title = "Official Timetable"
    ws.sheet_view.showGridLines = False

    HEADERS = [
        "Date", "Start Time", "End Time", "Duration",
        "Module Code", "Module Name", "Lecturer", "Location",
        "Session Status"
    ]
    num_cols = len(HEADERS)

    # Module colour palette — add more codes here as needed
    MODULE_COLORS = {
        "PUSL2022": "CCE5FF",  # soft blue
        "PUSL2021": "FFE5CC",  # soft orange
        "PUSL3123": "D4EDDA",  # soft green
        "PUSL3122": "F8D7DA",  # soft red
        "PUSL3124": "FFF3CD",  # soft yellow
        "PUSL2023": "E2CCFF",  # soft purple
        "PUSL2024": "CCF2FF",  # soft cyan
        "PUSL3121": "FFD6E0",  # soft pink
    }
    _DEFAULT_MODULE_FILL_HEX = "F0F0F0"  # neutral light grey fallback

    #  Metadata block (rows 1-4) 
    meta_rows = [
        (f"Official Academic Timetable — Batch {batch_id}", _TITLE_FONT),
        (f"Generated On: {datetime.now().strftime('%d %B %Y, %I:%M %p')}", _META_FONT),
        ("System Generated Document — University Academic Portal",           _META_FONT),
        ("",                                                                  None),
    ]
    for text, font in meta_rows:
        ws.append([text] + [""] * (num_cols - 1))
        row_idx = ws.max_row
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=num_cols)
        cell = ws.cell(row=row_idx, column=1)
        cell.alignment = _CENTER
        if font:
            cell.font = font

    #  Column header row 
    ws.append(HEADERS)
    header_row = ws.max_row
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font    = _HEADER_FONT
        cell.fill    = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border  = _HEAD_BORDER

    #  Data rows 
    for i, row in enumerate(rows):
        mod_code  = row["module_code"].strip().upper()
        mod_hex   = MODULE_COLORS.get(mod_code, _DEFAULT_MODULE_FILL_HEX)
        mod_fill  = PatternFill("solid", fgColor=mod_hex)
        mod_name_font = Font(name="Calibri", bold=True, size=10)  # Bold Module Name

        ws.append([
            row["date_fmt"],
            row["start_time"],
            row["end_time"],
            row["duration"],
            row["module_code"],
            row["module_name"],
            row["lecturer"],
            row["location"],
            row["status"],
        ])
        data_row = ws.max_row
        is_alt   = (i % 2 == 1)

        # Column positions in new layout (1-indexed):
        # 1=Date, 2=Start Time, 3=End Time, 4=Duration,
        # 5=Module Code, 6=Module Name, 7=Lecturer, 8=Location, 9=Session Status
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.border    = _CELL_BORDER
            cell.alignment = _CENTER if col_idx in (1, 2, 3, 4, 9) else _LEFT

            if col_idx == 9:
                # Status column — colour-coded badge
                cell.fill = _STATUS_FILLS.get(row["status"], _STATUS_FILLS["Unknown"])
                cell.font = _STATUS_FONTS.get(row["status"], _STATUS_FONTS["Unknown"])
                cell.alignment = _CENTER
            elif col_idx == 5:
                # Module Code — tinted by module palette
                cell.fill = mod_fill
            elif col_idx == 6:
                # Module Name — tinted + bold
                cell.fill = mod_fill
                cell.font = mod_name_font
            else:
                # Zebra striping for all other columns
                if is_alt:
                    cell.fill = _ALT_FILL

    #  Auto column widths 
    for col_idx, header in enumerate(HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in rows:
            vals = [
                row["date_fmt"], row["start_time"], row["end_time"], row["duration"],
                row["module_code"], row["module_name"], row["lecturer"],
                row["location"], row["status"],
            ]
            max_len = max(max_len, len(str(vals[col_idx - 1])))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze header rows
    ws.freeze_panes = f"A{header_row + 1}"

    #  SHEET 2 — Analytics Summary 
    ws2 = wb.create_sheet(title="Analytics Summary")
    ws2.sheet_view.showGridLines = False

    #  Title 
    ws2.append(["Timetable Analytics Summary"] + [""] * 3)
    ws2.merge_cells("A1:D1")
    title_cell = ws2["A1"]
    title_cell.font      = _TITLE_FONT
    title_cell.alignment = _CENTER

    ws2.append([f"Batch: {batch_id}"] + [""] * 3)
    ws2.merge_cells("A2:D2")
    ws2["A2"].font      = _META_FONT
    ws2["A2"].alignment = _CENTER

    ws2.append([f"Report Date: {datetime.now().strftime('%d %B %Y')}"] + [""] * 3)
    ws2.merge_cells("A3:D3")
    ws2["A3"].font      = _META_FONT
    ws2["A3"].alignment = _CENTER
    ws2.append([])

    # Total sessions
    ws2.append(["Total Scheduled Sessions", len(rows)] + [""] * 2)
    total_row = ws2.max_row
    ws2.cell(total_row, 1).font      = Font(name="Calibri", bold=True, size=11)
    ws2.cell(total_row, 2).font      = Font(name="Calibri", bold=True, size=11, color="1F3864")
    ws2.cell(total_row, 1).alignment = _LEFT
    ws2.cell(total_row, 2).alignment = _CENTER
    ws2.append([])

    # Status breakdown
    from collections import Counter
    status_counts = Counter(r["status"] for r in rows)
    ws2.append(["Session Status Breakdown"] + [""] * 3)
    ws2.merge_cells(f"A{ws2.max_row}:D{ws2.max_row}")
    ws2.cell(ws2.max_row, 1).font      = Font(name="Calibri", bold=True, size=11, color=_DARK_BLUE)
    ws2.cell(ws2.max_row, 1).alignment = _CENTER

    status_headers = ["Status", "Count"]
    ws2.append(status_headers + [""] * 2)
    sh_row = ws2.max_row
    for col_idx, h in enumerate(status_headers, start=1):
        c = ws2.cell(sh_row, col_idx)
        c.font = _HEADER_FONT; c.fill = _HEADER_FILL
        c.alignment = _CENTER; c.border = _HEAD_BORDER

    for s_label, s_count in status_counts.items():
        ws2.append([s_label, s_count, "", ""])
        r = ws2.max_row
        ws2.cell(r, 1).fill      = _STATUS_FILLS.get(s_label, _STATUS_FILLS["Unknown"])
        ws2.cell(r, 1).font      = _STATUS_FONTS.get(s_label, _STATUS_FONTS["Unknown"])
        ws2.cell(r, 1).alignment = _CENTER
        ws2.cell(r, 2).alignment = _CENTER
        ws2.cell(r, 2).font      = Font(name="Calibri", bold=True)
        for col_idx in range(1, 3):
            ws2.cell(r, col_idx).border = _CELL_BORDER
    ws2.append([])

    # Per-lecturer hours summary
    from collections import defaultdict
    lecturer_hours: dict[str, float] = defaultdict(float)
    for r in rows:
        lecturer_hours[r["lecturer"]] += r["duration_h"]

    ws2.append(["Total Lecture Hours per Lecturer"] + [""] * 3)
    ws2.merge_cells(f"A{ws2.max_row}:D{ws2.max_row}")
    ws2.cell(ws2.max_row, 1).font      = Font(name="Calibri", bold=True, size=11, color=_DARK_BLUE)
    ws2.cell(ws2.max_row, 1).alignment = _CENTER

    lec_headers = ["Lecturer", "Total Sessions", "Total Hours", "Avg. Session Length"]
    ws2.append(lec_headers)
    lh_row = ws2.max_row
    for col_idx, h in enumerate(lec_headers, start=1):
        c = ws2.cell(lh_row, col_idx)
        c.font = _HEADER_FONT; c.fill = _HEADER_FILL
        c.alignment = _CENTER; c.border = _HEAD_BORDER

    for i, (lec, total_h) in enumerate(
        sorted(lecturer_hours.items(), key=lambda x: x[1], reverse=True)
    ):
        sessions_for_lec = sum(1 for r in rows if r["lecturer"] == lec)
        avg_h = total_h / sessions_for_lec if sessions_for_lec else 0
        ws2.append([
            lec,
            sessions_for_lec,
            _fmt_duration(total_h),
            _fmt_duration(avg_h),
        ])
        r_idx = ws2.max_row
        is_alt = (i % 2 == 1)
        for col_idx in range(1, 5):
            c = ws2.cell(r_idx, col_idx)
            c.border    = _CELL_BORDER
            c.alignment = _CENTER if col_idx > 1 else _LEFT
            if is_alt:
                c.fill = _ALT_FILL

    # Auto-widths for Sheet 2
    for col_idx in range(1, 5):
        col_letter = get_column_letter(col_idx)
        max_len = len(lec_headers[col_idx - 1])
        for row_vals in ws2.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            max_len = max(max_len, len(str(row_vals[0] or "")))
        ws2.column_dimensions[col_letter].width = min(max_len + 4, 40)

    #  Stream workbook buffer 
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_batch = batch_id.replace(" ", "_").replace("/", "-")
    filename   = f"Official_Timetable_{safe_batch}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/template")
async def download_timetable_template():
    """
    Generates and serves a professional, styled Excel upload template.
    Sheet 1: 'Timetable Template' — headers + sample row + blank data rows.
    Sheet 2: 'Instructions'       — numbered formatting rules for Admins.
    """
    wb = openpyxl.Workbook()

    # ── Shared style tokens ────────────────────────────────────────────────────
    TEMPLATE_HEADERS = [
        "Date", "Day", "Start Time", "End Time",
        "Module Code", "Module Name", "Lecturer", "Location"
    ]
    COL_WIDTHS = [16, 12, 14, 14, 16, 34, 28, 16]  # aligned to each header

    # Header: dark green bg, bold white text
    HDR_FILL = PatternFill("solid", fgColor="1A5276")
    HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

    # Sample row: soft mint bg, italic font to signal "example only"
    SAMPLE_FILL = PatternFill("solid", fgColor="D5F5E3")
    SAMPLE_FONT = Font(name="Calibri", italic=True, color="1D6A39", size=10)

    # Blank data rows: very light grey alternating tint
    DATA_FILL_ALT = PatternFill("solid", fgColor="F4F6F7")

    T_CENTER = Alignment(horizontal="center", vertical="center")
    T_LEFT   = Alignment(horizontal="left",   vertical="center")

    T_THIN   = Side(style="thin",   color="BDC3C7")
    T_THICK  = Side(style="medium", color="1A5276")
    CELL_BDR = Border(left=T_THIN, right=T_THIN, top=T_THIN, bottom=T_THIN)
    HEAD_BDR = Border(left=T_THICK, right=T_THICK, top=T_THICK, bottom=T_THICK)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Timetable Template
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Timetable Template"
    ws.sheet_view.showGridLines = False

    # ── Title block (rows 1–3) ────────────────────────────────────────────────
    num_cols = len(TEMPLATE_HEADERS)
    title_meta = [
        ("Timetable Upload Template",
         Font(name="Calibri", bold=True, size=14, color="1A5276")),
        (f"Generated: {datetime.now().strftime('%d %B %Y')} — University Academic Portal",
         Font(name="Calibri", italic=True, size=10, color="7F8C8D")),
        ("⚠  Fill from Row 5 onwards. Do NOT modify header names or column order.",
         Font(name="Calibri", bold=True, size=10, color="C0392B")),
        ("", None),
    ]
    for text, font in title_meta:
        ws.append([text] + [""] * (num_cols - 1))
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
        ws.cell(r, 1).alignment = T_CENTER
        if font:
            ws.cell(r, 1).font = font

    # ── Column headers (row 5) ────────────────────────────────────────────────
    ws.append(TEMPLATE_HEADERS)
    hdr_row = ws.max_row
    for col_idx in range(1, num_cols + 1):
        c = ws.cell(hdr_row, col_idx)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = T_CENTER
        c.border    = HEAD_BDR

    # ── Sample row (row 6) ────────────────────────────────────────────────────
    SAMPLE_ROW = [
        "2026-05-20", "Wednesday", "09:00 AM", "12:00 PM",
        "PUSL3190", "Computing Project",
        "Ms. Sanuli Weerasinghe", "Hall A"
    ]
    ws.append(SAMPLE_ROW)
    sample_row = ws.max_row
    for col_idx in range(1, num_cols + 1):
        c = ws.cell(sample_row, col_idx)
        c.font      = SAMPLE_FONT
        c.fill      = SAMPLE_FILL
        c.alignment = T_CENTER if col_idx in (1, 2, 3, 4, 5) else T_LEFT
        c.border    = CELL_BDR

    # Add a note label in the column immediately after the last data column
    label_cell = ws.cell(sample_row, num_cols + 1)
    label_cell.value     = "← Sample Row (delete before uploading)"
    label_cell.font      = Font(name="Calibri", italic=True, color="7F8C8D", size=9)
    label_cell.alignment = T_LEFT

    # ── Blank data entry rows (rows 7–26, 20 rows) ────────────────────────────
    for row_idx in range(20):
        ws.append([""] * num_cols)
        r = ws.max_row
        for col_idx in range(1, num_cols + 1):
            c = ws.cell(r, col_idx)
            c.border    = CELL_BDR
            c.alignment = T_CENTER if col_idx in (1, 2, 3, 4, 5) else T_LEFT
            if row_idx % 2 == 1:
                c.fill = DATA_FILL_ALT

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[hdr_row].height = 22
    ws.freeze_panes = f"A{hdr_row + 1}"  # Freeze above data

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Instructions
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="Instructions")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 72

    # Sheet title
    ws2.append(["", "📋  Timetable Upload — Formatting Instructions"])
    title_r = ws2.max_row
    ws2.cell(title_r, 2).font      = Font(name="Calibri", bold=True, size=14, color="1A5276")
    ws2.cell(title_r, 2).alignment = T_LEFT
    ws2.append([])

    RULES = [
        ("Date Column",
         "Use YYYY-MM-DD format exclusively (e.g., 2026-05-20). Excel date objects are also accepted."),
        ("Time Columns",
         "Use HH:MM AM/PM format for both Start Time and End Time (e.g., 09:00 AM, 12:00 PM)."),
        ("Module Code",
         "Must exactly match a registered Module Code in the system (e.g., PUSL3190). Case-insensitive."),
        ("Module Name",
         "Must match the official module name registered in the system. Leading/trailing spaces are stripped."),
        ("Lecturer",
         "Must exactly match the lecturer's registered full name. Use 'TBA' if the lecturer is not yet assigned."),
        ("Location",
         "Enter the room/hall designation (e.g., Hall A, Lab 01). Use 'TBD' if the location is unknown."),
        ("Column Order",
         "Do NOT rearrange, rename, or delete any column headers. The system validates the template structure."),
        ("Sample Row",
         "Row 6 in the Template sheet is a sample row for guidance. Delete it before uploading."),
        ("Empty Rows",
         "Empty rows within the data range are safely ignored during extraction."),
        ("File Format",
         "Save and upload as .xlsx (Excel Workbook). CSV format is not accepted for this template."),
    ]

    # Section header for rules table
    ws2.append(["", "Rule"])
    ws2.append(["#", "Field", "Guideline"][0:1] + ["Field — Guideline"])
    sec_r = ws2.max_row
    for col_idx in (1, 2):
        c = ws2.cell(sec_r, col_idx)
        c.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor="1A5276")
        c.alignment = T_CENTER if col_idx == 1 else T_LEFT
        c.border    = HEAD_BDR

    for i, (field, rule) in enumerate(RULES, start=1):
        ws2.append([str(i), f"{field}:  {rule}"])
        r = ws2.max_row
        ws2.cell(r, 1).font      = Font(name="Calibri", bold=True, size=10, color="1A5276")
        ws2.cell(r, 1).alignment = T_CENTER
        ws2.cell(r, 2).font      = Font(name="Calibri", size=10)
        ws2.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws2.row_dimensions[r].height = 28
        for col_idx in (1, 2):
            ws2.cell(r, col_idx).border = CELL_BDR
            if i % 2 == 0:
                ws2.cell(r, col_idx).fill = PatternFill("solid", fgColor="EBF5FB")

    # ── Stream buffer ─────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Timetable_Upload_Template.xlsx"'},
    )

# Mapping for weekday offsets
DAY_OFFSET = {
    "Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3,
    "Friday": 4, "Saturday": 5, "Sunday": 6
}

@router.post("/extract")
async def extract_standard_timetable(
    file: UploadFile = File(...),
    faculty: str = Form(...),
    department: str = Form(...),
    batch: str = Form(...),
    semester: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    Standardized ETL endpoint for the flat Excel template.
    Ensures data integrity through strict validation and lecturer existence checks.
    """
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The extraction pipeline currently only supports Excel files (.xlsx, .xls)."
        )

    try:
        contents = await file.read()

        # Auto-detect the real header row (handles styled templates that have
        # a title block above the actual column names).
        # Strategy: read without a header, then scan the first 10 rows for one
        # that contains the word "date" (case-insensitive) — that is our header.
        raw_df = pd.read_excel(io.BytesIO(contents), header=None)
        header_row_idx = 0  # default: first row is the header (plain flat file)
        for i, row in raw_df.head(10).iterrows():
            row_vals = [str(v).strip().lower() for v in row.values if pd.notna(v)]
            if "date" in row_vals:
                header_row_idx = i
                break

        df = pd.read_excel(io.BytesIO(contents), header=header_row_idx)

        # 1. Clean column names to avoid trailing space issues
        df.columns = df.columns.astype(str).str.strip().str.title()

        # 2. Strict Column Validation
        required_cols = ['Date', 'Day', 'Start Time', 'End Time', 'Module Code', 'Module Name', 'Lecturer', 'Location']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid template format. Missing required columns: {', '.join(missing_cols)}. Please use the Standard Template."
            )

        validation_errors = []

        # Check each row for valid mapping
        for index, row in df.iterrows():
            # Skip validation if critical fields are missing (Extraction loop will skip these later)
            if pd.isna(row['Module Code']) or pd.isna(row['Lecturer']):
                continue
                
            raw_lec_name = str(row['Lecturer']).strip()
            mod_code = str(row['Module Code']).strip().upper()

            # Handle TBA explicitly
            if raw_lec_name.upper() == "TBA":
                continue

            # Check 1: Is Lecturer in the system? (Bulletproof Case-Insensitive Query)
            db_lecturer = db.query(Lecturer).filter(
                func.lower(func.trim(Lecturer.name)) == raw_lec_name.lower()
            ).first()

            if not db_lecturer:
                err = f"Lecturer '{raw_lec_name}' (Row {index+2}) is not registered in the system."
                if err not in validation_errors:
                    validation_errors.append(err)
                continue
            
            # Check 2: Is Lecturer assigned to this module?
            # Check the comma-separated assigned_subjects field
            assigned_list = []
            if db_lecturer.assigned_subjects:
                # Assuming comma-separated: "PUSL2022, PUSL2021"
                assigned_list = [s.strip().upper() for s in db_lecturer.assigned_subjects.split(",")]

            if mod_code not in assigned_list:
                err = f"Validation Error (Row {index+2}): '{db_lecturer.name}' is not officially assigned to teach '{mod_code}'."
                if err not in validation_errors:
                    validation_errors.append(err)

        if validation_errors:
            raise HTTPException(
                status_code=400,
                detail={"message": "Validation Failed", "errors": validation_errors}
            )

        # 4. Extract Data
        extracted_data = []
        for _, row in df.iterrows():
            # Skip completely empty rows or rows missing heart data
            if pd.isna(row['Module Code']) or pd.isna(row['Date']):
                continue
                
            # Handle Pandas datetime objects gracefully for the 'Date' column
            date_val = str(row['Date']).split(' ')[0] if pd.notna(row['Date']) else ""
            
            # Auto-resolve degree from database using module code
            mod_code = str(row['Module Code']).strip().upper()
            db_module = db.query(Module).filter(Module.module_code == mod_code).first()
            resolved_degree = db_module.degree if db_module else "Unknown"

            extracted_data.append({
                "date": date_val,
                "day": str(row['Day']).strip(),
                "time": f"{str(row['Start Time']).strip()} - {str(row['End Time']).strip()}",
                "module_code": mod_code,
                "module": str(row['Module Name']).strip(),
                "lecturer": str(row['Lecturer']).strip(),
                "location": str(row['Location']).strip(),
                "faculty": faculty,
                "department": department,
                "batch": batch,
                "semester": semester
            })

        return {
            "status": "success",
            "message": "Timetable extracted successfully",
            "total_records": len(extracted_data),
            "full_data": extracted_data # Return everything for final sync
        }

    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Extraction failed: {str(e)}"
        )

@router.post("/upload", status_code=status.HTTP_201_CREATED)
async def upload_timetable(
    batch_id: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Parses a 2D Matrix Timetable (Excel/CSV) and inserts it into the DB.
    Expects Days/Dates on X-axis (headers) and Time on Y-axis (first column).
    """
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file format. Only .csv, .xlsx, .xls are supported."
        )

    try:
        contents = await file.read()
        
        # Parse into Pandas DataFrame
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents))
            
        # Clean dataframe: drop completely empty rows and columns
        df.dropna(how='all', inplace=True)
        df.dropna(axis=1, how='all', inplace=True)
        
        if df.empty or len(df.columns) < 2:
            raise ValueError("The uploaded file does not contain a valid matrix grid.")

        # Identify the Time column (usually the very first column in a 2D matrix timetable)
        time_col_name = df.columns[0]
        
        # The remaining columns represent Dates or Days (e.g., "Monday", "2024-03-25")
        date_columns = df.columns[1:]
        
        entries_to_insert = []
        
        # Iterate over the grid
        for idx, row in df.iterrows():
            time_str = str(row[time_col_name]).strip()
            if not time_str or time_str.lower() == 'nan':
                continue
                
            # Attempt to split "09:00 - 10:00" into start and end time
            parts = [p.strip() for p in time_str.split('-')]
            if len(parts) == 2:
                start_time, end_time = parts[0], parts[1]
            else:
                # If cannot split, fallback to treating the whole string as start_time
                start_time = time_str
                end_time = ""  # We can't definitively know end_time in this case
                
            for date_col in date_columns:
                cell_val = str(row[date_col]).strip()
                if not cell_val or cell_val.lower() == 'nan':
                    continue # Empty cell means no class at this time/date
                    
                # We found a module for this day and time!
                entries_to_insert.append(
                    Timetable(
                        batch_id=batch_id,
                        module_code=cell_val,
                        date=str(date_col).strip(),
                        start_time=start_time,
                        end_time=end_time
                    )
                )
                
        if not entries_to_insert:
            raise ValueError("Could not find any valid classes in the provided matrix.")

        # Batch insert into Database
        db.add_all(entries_to_insert)
        db.commit()
        
        log_audit_action(db, "Timetable Upload", f"Uploaded timetable for batch {batch_id} with {len(entries_to_insert)} entries.")
        
        return {
            "message": "Timetable processed and saved successfully.",
            "entries_added": len(entries_to_insert)
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to parse timetable: {str(e)}"
        )

class TimetableRecord(BaseModel):
    date: str
    day: str
    time: str
    module_code: str
    module_name: str = ""
    lecturer: str = ""
    location: str = ""
    batch: str = ""
    model_config = {"extra": "ignore"}

class SyncPayload(BaseModel):
    batch_id: str
    faculty: str
    department: str
    semester: str
    file_name: str
    records: List[TimetableRecord]

@router.post("/sync")
async def sync_timetable(payload: SyncPayload, db: Session = Depends(get_db)):
    """
    Persists reviewed timetable records into the database.
    Splits the 'time' range into start_time and end_time.
    """
    try:
        # Check if timetable already exists for this batch and semester to prevent duplicates
        existing_record = db.query(Timetable).filter(
            Timetable.batch_id == payload.batch_id,
            Timetable.semester == payload.semester
        ).first()

        if existing_record:
            raise HTTPException(
                status_code=400,
                detail=f"Batch '{payload.batch_id}' ({payload.semester}) already exists. Delete the old one first."
            )

        # Transactional insert
        for record in payload.records:
            # Split "09:00 AM - 11:00 AM" or "09:00 - 11:00"
            time_parts = [p.strip() for p in record.time.split("-")]
            start_t = time_parts[0] if len(time_parts) > 0 else record.time
            end_t = time_parts[1] if len(time_parts) > 1 else ""

            new_entry = Timetable(
                batch_id=payload.batch_id,
                faculty=payload.faculty,
                department=payload.department,
                semester=payload.semester,
                file_name=payload.file_name,
                module_code=record.module_code,
                module_name=record.module_name,
                date=record.date,
                start_time=start_t,
                end_time=end_t,
                lecturer=record.lecturer,
                location=record.location
            )
            db.add(new_entry)
        
        db.commit()
        log_audit_action(db, "System Operations", f"Synced {len(payload.records)} records for batch {payload.batch_id} (File: {payload.file_name})")
        
        return {"status": "success", "message": f"Successfully synced {len(payload.records)} records."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")

@router.get("/recent")
async def get_recent_uploads(db: Session = Depends(get_db)):
    """
    Bulletproof retrieval of unique batches. 
    Groups uniquely in Python to avoid SQLAlchemy/DB-dialect serialization and grouping oddities.
    """
    try:
        # 1. Simple, safe query: Get all records ordered by newest first
        records = db.query(Timetable.batch_id, Timetable.created_at, Timetable.file_name).order_by(Timetable.created_at.desc()).all()

        # 2. Group uniquely in Python
        seen_batches = set()
        recent_uploads = []

        for row in records:
            if row.batch_id not in seen_batches:
                seen_batches.add(row.batch_id)
                
                # 3. Bulletproof string conversion
                safe_date = row.created_at.strftime("%b %d, %Y") if hasattr(row.created_at, 'strftime') else str(row.created_at)
                
                recent_uploads.append({
                    "id": row.batch_id,
                    "name": row.file_name or f"Batch_{row.batch_id}_Schedule.xlsx",
                    "batch": row.batch_id,
                    "date": safe_date,
                    "status": "Success"
                })

        return recent_uploads
    
    except Exception as e:
        print(f"RECENT UPLOADS ERROR: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch recent uploads: {str(e)}")

@router.get("/batch/{batch_id}")
async def get_timetable_by_batch(batch_id: str, db: Session = Depends(get_db)):
    """
    Fetches all timetable records for a specific batch.
    Used by the frontend View Modal.
    """
    try:
        # Fetch all records for the given batch_id, ordered by date and time
        records = db.query(Timetable).filter(
            Timetable.batch_id == batch_id
        ).order_by(Timetable.date.asc(), Timetable.start_time.asc()).all()
        
        if not records:
            raise HTTPException(status_code=404, detail=f"No records found for batch {batch_id}")
            
        # Serialize the records
        formatted_records = []
        for rec in records:
            # We use getattr in case these columns are newly added but might be null in old records
            formatted_records.append({
                "id": rec.id,
                "module_code": rec.module_code,
                "module_name": getattr(rec, 'module_name', rec.module_code), 
                "date": rec.date,
                "start_time": rec.start_time,
                "end_time": rec.end_time,
                "lecturer": rec.lecturer,
                "faculty": rec.faculty,
                "department": rec.department,
                "semester": rec.semester
            })
            
        return formatted_records
    
    except HTTPException:
        raise
    except Exception as e:
        print(f"VIEW BATCH ERROR: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch batch records: {str(e)}")

@router.delete("/batch/{batch_id}")
async def delete_timetable_batch(batch_id: str, db: Session = Depends(get_db)):
    """
    Deletes all timetable records for a specific academic batch.
    """
    try:
        deleted_count = db.query(Timetable).filter(Timetable.batch_id == batch_id).delete()
        db.commit()
        
        log_audit_action(db, "System Operations", f"Deleted timetable batch {batch_id} ({deleted_count} records purged)")
        
        return {"status": "success", "message": f"Deleted {deleted_count} records for batch {batch_id}."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Deletion failed: {str(e)}")
@router.post("/{session_id}/start")
async def start_timetable_session(session_id: int, db: Session = Depends(get_db)):
    """Sets a timetable session as active/live."""
    # 1. Update the high-level Timetable status (Used by Dashboards)
    session = db.query(Timetable).filter(Timetable.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found in timetable")
    session.is_live = True

    # 2. Synchronize with the active ClassSession tracker (Used by Lecturer Monitoring)
    # The session_id here maps to what the monitoring component expects
    class_session = db.query(models.ClassSession).filter(models.ClassSession.id == session_id).first()
    if class_session:
        class_session.status = "Live"

    db.commit()
    return {"status": "success", "message": f"Session {session_id} is now LIVE."}

@router.post("/{session_id}/stop")
async def stop_timetable_session(session_id: int, db: Session = Depends(get_db)):
    """Ends a live timetable session and marks it as Completed."""
    # 1. Update the high-level Timetable status
    session = db.query(Timetable).filter(Timetable.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found in timetable")
    session.is_live = False
    session.status = "Completed"  # CRITICAL: Write to the new status column so hybrid logic can detect it
    # NOTE: Do NOT clear cover_requested here. The flag must be preserved after
    # completion so the frontend can correctly show "Covered by Admin" vs "Session Completed".

    # 2. Synchronize with the active ClassSession tracker
    class_session = db.query(models.ClassSession).filter(models.ClassSession.id == session_id).first()
    if class_session:
        class_session.status = "Completed"  # Match the string the frontend checks

    db.commit()
    return {"status": "success", "message": f"Session {session_id} is now COMPLETED."}
class CoverRequest(BaseModel):
    reason: str

@router.post("/{session_id}/request_cover")
async def request_cover(session_id: int, payload: CoverRequest, db: Session = Depends(get_db)):
    """Allows a lecturer to request an admin to cover their class."""
    session = db.query(Timetable).filter(Timetable.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    session.cover_requested = True
    session.cover_reason = payload.reason
    db.commit()
    
    return {"status": "success", "message": "Cover request submitted successfully."}
